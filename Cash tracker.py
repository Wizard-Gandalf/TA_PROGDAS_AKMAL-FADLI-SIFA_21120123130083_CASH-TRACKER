import tkinter as tk
from tkinter import ttk, messagebox
import threading
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
import os

FILE_EXCEL = "data_pengeluaran.xlsx"
refresh_lock = False  # mencegah loop refresh tak berujung


# ---------------------------------------------------------
#   MEMBUAT FILE EXCEL JIKA BELUM ADA
# ---------------------------------------------------------
def init_excel():
    if not os.path.exists(FILE_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pengeluaran"
        ws.append(["Nama Pengeluaran", "Nominal"])
        wb.save(FILE_EXCEL)


# ---------------------------------------------------------
#   FUNGSI MEMUAT DATA DARI EXCEL KE GUI
# ---------------------------------------------------------
def load_from_excel():
    global refresh_lock
    if refresh_lock:
        return

    try:
        wb = load_workbook(FILE_EXCEL)
        ws = wb.active

        tabel.delete(*tabel.get_children())
        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            nama, nominal = row
            tabel.insert("", "end", values=(nama, nominal))
            total += float(nominal)

        label_total.config(text=f"Total Pengeluaran: Rp {total:,.2f}")

    except Exception as e:
        print("Gagal memuat Excel:", e)


# ---------------------------------------------------------
#   FUNGSI MENYIMPAN DATA DARI GUI KE EXCEL
# ---------------------------------------------------------
def save_to_excel():
    global refresh_lock
    refresh_lock = True  # mencegah watchdog memicu update GUI dari perubahan Excel

    wb = Workbook()
    ws = wb.active
    ws.append(["Nama Pengeluaran", "Nominal"])

    for row_id in tabel.get_children():
        nama, nominal = tabel.item(row_id)["values"]
        ws.append([nama, float(nominal)])

    wb.save(FILE_EXCEL)

    refresh_lock = False


# ---------------------------------------------------------
#   FUNGSI TAMBAH PENGELUARAN
# ---------------------------------------------------------
def tambah_pengeluaran():
    nama = entry_nama.get().strip()
    nominal = entry_nominal.get().strip()

    if not nama or not nominal:
        messagebox.showwarning("Peringatan", "Nama dan nominal harus diisi.")
        return

    try:
        nominal = float(nominal)
    except:
        messagebox.showerror("Error", "Nominal harus angka.")
        return

    tabel.insert("", "end", values=(nama, nominal))

    hitung_total()
    save_to_excel()

    entry_nama.delete(0, tk.END)
    entry_nominal.delete(0, tk.END)


# ---------------------------------------------------------
#   HITUNG ULANG TOTAL
# ---------------------------------------------------------
def hitung_total():
    total = 0
    for row in tabel.get_children():
        total += float(tabel.item(row)["values"][1])

    label_total.config(text=f"Total Pengeluaran: Rp {total:,.2f}")


# ---------------------------------------------------------
#   EDIT DATA TERPILIH
# ---------------------------------------------------------
def edit_data():
    selected = tabel.selection()
    if not selected:
        messagebox.showwarning("Peringatan", "Pilih data terlebih dahulu.")
        return

    row_id = selected[0]
    nama_lama, nominal_lama = tabel.item(row_id)["values"]

    popup = tk.Toplevel(root)
    popup.title("Edit Data")
    popup.geometry("300x150")

    tk.Label(popup, text="Nama:").pack()
    nama_entry = tk.Entry(popup)
    nama_entry.insert(0, nama_lama)
    nama_entry.pack()

    tk.Label(popup, text="Nominal:").pack()
    nominal_entry = tk.Entry(popup)
    nominal_entry.insert(0, nominal_lama)
    nominal_entry.pack()

    def simpan_edit():
        try:
            nominal_baru = float(nominal_entry.get())
        except:
            messagebox.showerror("Error", "Nominal harus angka.")
            return

        nama_baru = nama_entry.get().strip()
        tabel.item(row_id, values=(nama_baru, nominal_baru))
        hitung_total()
        save_to_excel()
        popup.destroy()

    tk.Button(popup, text="Simpan", command=simpan_edit).pack(pady=10)


# ---------------------------------------------------------
#   HAPUS DATA
# ---------------------------------------------------------
def hapus_data():
    selected = tabel.selection()
    if not selected:
        messagebox.showwarning("Peringatan", "Pilih data yang ingin dihapus.")
        return

    tabel.delete(selected[0])
    hitung_total()
    save_to_excel()


# ---------------------------------------------------------
#   WATCHDOG UNTUK MENGAWASI PERUBAHAN FILE EXCEL
# ---------------------------------------------------------
class ExcelEventHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if FILE_EXCEL in event.src_path:
            load_from_excel()


def start_watchdog():
    event_handler = ExcelEventHandler()
    observer = Observer()
    observer.schedule(event_handler, ".", recursive=False)
    observer.start()


# ---------------------------------------------------------
#   GUI TKINTER
# ---------------------------------------------------------
root = tk.Tk()
root.title("Pengeluaran Harian (Real-Time Excel)")
root.geometry("600x500")

judul = tk.Label(root, text="Aplikasi Pengeluaran Harian – Sinkron Excel Real-Time",
                  font=("Segoe UI", 14, "bold"))
judul.pack(pady=10)

frame_input = tk.Frame(root)
frame_input.pack(pady=5)

tk.Label(frame_input, text="Nama:").grid(row=0, column=0)
entry_nama = tk.Entry(frame_input)
entry_nama.grid(row=0, column=1, padx=5)

tk.Label(frame_input, text="Nominal:").grid(row=1, column=0)
entry_nominal = tk.Entry(frame_input)
entry_nominal.grid(row=1, column=1, padx=5)

btn_tambah = tk.Button(root, text="Tambah", command=tambah_pengeluaran)
btn_tambah.pack(pady=5)

tabel = ttk.Treeview(root, columns=("nama", "nominal"), show="headings")
tabel.heading("nama", text="Nama Pengeluaran")
tabel.heading("nominal", text="Nominal")
tabel.pack(pady=10)

btn_edit = tk.Button(root, text="Edit", command=edit_data)
btn_edit.pack()

btn_hapus = tk.Button(root, text="Hapus", command=hapus_data)
btn_hapus.pack(pady=5)

label_total = tk.Label(root, text="Total Pengeluaran: Rp 0", font=("Segoe UI", 12, "bold"))
label_total.pack(pady=10)


# ---------------------------------------------------------
#   MULAI PROGRAM
# ---------------------------------------------------------
init_excel()
load_from_excel()

threading.Thread(target=start_watchdog, daemon=True).start()

root.mainloop()
